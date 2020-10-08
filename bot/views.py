# ------------------------------ Developer - Ayaz Saiyed M.
# --------------- September 2020

from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import Http404
from pydialogflow_fulfillment import DialogflowResponse, DialogflowRequest, SimpleResponse, Suggestions, LinkOutSuggestion
from django.http import HttpResponse, JsonResponse
from library.df_response_lib import *
from library.facebook_template_lib import *
import json
import requests
# import facebook
import time
import random
from .models import UsersDetails
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

from datetime import datetime
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 

fromaddr = "saiyedayaz9@gmail.com"
toaddr = "saiyedayaz9@gmail.com"

    # instance of MIMEMultipart 
msg = MIMEMultipart() 
    # storing the senders email address 
msg['From'] = fromaddr 
    # storing the receivers email address 
msg['To'] = toaddr 
    # storing the subject 
msg['Subject'] = "Customer Enquiry"

    # string to store the body of the mail 
body = "Attached Excel Sheet"
msg.attach(MIMEText(body, 'plain')) 

    # open the file to be sent 

    # creates SMTP session 
s = smtplib.SMTP('smtp.gmail.com', 587) 

    # start TLS for security 
s.starttls() 
    # Authentication 
s.login(fromaddr, "officialhearthackergmail7777") 

    # Converts the Multipart msg into a string 


def homepage(request):
    return render(request,'./bot/index.html')



@csrf_exempt
def index_function(request):
    xdate = datetime.now().strftime('%Y-%m-%d')
    Issuedate = xdate
    if request.method == "POST":

        print("Method ",request.method)

        if request.body:
            req = json.loads(request.body)
            dialogflow_request = DialogflowRequest(request.body)
            action = req.get('queryResult').get('action') 
            print("Action is ",action)
            



            if dialogflow_request.get_intent_displayName() == "AskingServices - LaptopIssueFetch - Final":
                print("First Intent - Laptop ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Perfecto, uno de nuestros 👨 ejecutivos se pondrá en contacto contigo en breve para solucionar el problema relacionado con tu portátil. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "problema de la computadora portátil"
                customerName = req.get('queryResult').get('parameters').get('username')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()
                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]


                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                # for column_cells in sheet.columns:
                #     # length = max(len("fewopfewpfrewprjkeddwioprfdsfsdfdsffsdfsdsfdsfdsfdsfdsfjewprjkeddwioprfdsfsdfdsffsdfsdsfdsfdsfdsfdsfjewprjkeddwioprfdsfsdfdsffsdfsdsfdsfdsfdsfdsfjewprjkeddwioprfdsfsdfdsffsdfsdsfdsfdsfdsfdsfjewpfdssdsdasdadrwpfrjipewr") for cell in column_cells)
                #     sheet.column_dimensions[column_cells[0].column_letter].width = int(20)

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length
                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                print(" Email Has Been Sent ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                # yy = sendEmail(request)
                # import time
                # time.sleep(0.5)
                response = dialogflow_response.get_final_response()


            if dialogflow_request.get_intent_displayName() == "AskingServices - ComputerIssueFetch - Final":
                print("Second Intent - Computer ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Perfecto, uno de nuestros 👨 ejecutivos se comunicará contigo en breve para solucionar tu problema informático. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "problema de la computadora"
                customerName = req.get('queryResult').get('parameters').get('username')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()

                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]


                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length


                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()


            if dialogflow_request.get_intent_displayName() == "AskingServices - Software":
                print(" Third Intent - Software ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Perfecto, uno de nuestros 👨 ejecutivos se comunicará con usted en breve para resolver su problema relacionado con el software. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "Problema de software"
                customerName = req.get('queryResult').get('parameters').get('name')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()

                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]

                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length


                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()



            if dialogflow_request.get_intent_displayName() == "AskingServices - Puntos de venta":
                print("Fourth Intent - Sales ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Genial, uno de nuestros 👨 ejecutivos se pondrá en contacto contigo en breve para discutir las ventas en detalle. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "Puntos de venta Issue"
                customerName = req.get('queryResult').get('parameters').get('name')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()

                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]

                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length


                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()




            if dialogflow_request.get_intent_displayName() == "AskingServices - Redes":
                print("Fifth Intent - Redes ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Seguro, uno de nuestros expertos en nuestra red se pondrá en contacto contigo para solucionar tu problema 👍 Que tengas un buen día 🤝 Gracias 💫 ")
                customerNeed = "Problema de red"
                customerName = req.get('queryResult').get('parameters').get('username')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()

                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                if not customerName:
                    customerName=="None"
                if not customerEmail:
                    customerEmail=="None"
                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]

                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length


                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()



            if dialogflow_request.get_intent_displayName() == "AskingServices - Venta de Equip de Computo":
                print("Sixth Intent - Venta de Equip de Computo ")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("De acuerdo, nos comunicaremos con usted en breve para discutir en detalle sobre el equipo que se venderá. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "Venta de Equip de Computo Issue"
                customerName = req.get('queryResult').get('parameters').get('name')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()

                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]

                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")


                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length


                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()




            if dialogflow_request.get_intent_displayName() == "AskingServices - Otro - Final":
                print("Seventh Intent - Otro")
                # print("City Selection") 
                QuerySet = req.get('queryResult').get('parameters')
                print("QuerySet ",QuerySet)
                dialogflow_response = DialogflowResponse("Pues uno de nuestros ejecutivos de soporte se comunicará contigo en breve para brindarte el mejor servicio👨. Que tengas un buen día 🤝 Gracias 💫")
                customerNeed = "Otro"
                customerName = req.get('queryResult').get('parameters').get('name')
                customerEmail = req.get('queryResult').get('parameters').get('email')
                response = dialogflow_response.get_final_response()


                print("")
                print(" Customer name is ",customerName)
                print(" Customer email is ",customerEmail)
                print(" Customer Issue is ",customerNeed)
                print("")

                #Excel stuff

                workbook = Workbook()
                workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
                sheet = workbook.active

                # Details of User
                rows = [
                            [customerName,customerNeed,customerEmail,Issuedate]
                       ]


                for row in rows:
                    sheet.append(row)

                workbook.save(filename="SoporteMoreliaCustomers.xlsx")

                for column_cells in sheet.columns:
                    length = max(len(cell.value) for cell in column_cells)
                    sheet.column_dimensions[column_cells[0].column_letter].width = length




                workbook.save(filename="SoporteMoreliaCustomers.xlsx")
                print(" Data Saved ")
                action = req.get('queryResult').get('action')   
                x = action
                filename = "SoporteMoreliaCustomers.xlsx"
                attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 
                p = MIMEBase('application', 'octet-stream') 
                p.set_payload((attachment).read()) 
                encoders.encode_base64(p) 
                p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
                    # attach the instance 'p' to instance 'msg' 
                msg.attach(p)
                text = msg.as_string() 

                s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
                s.quit()
                response = dialogflow_response.get_final_response()



        else :
            response = {
                "error" : "1",
                "message" : "An error occurred."
            }
        print("sending baby")
        
        return HttpResponse(response, content_type='application/json; charset=utf-8')
        # return HttpResponse("Yudiz Team")

    else:
        # print("3")

        raise Http404()


def get_user_first_name(user_id):
    '''
        Retrieves user first name using Facebook Graph API for a user with user_id
    '''
    user = graph.get_object(id=str(user_id))
    if user.get('first_name'):
        return user.get('first_name')
    else:
        return False


# def uploadtoDrive(request):
#     import json
#     import requests

#     filedirectory = '###'
#     filename = '###'
#     folderid = '###'
#     access_token = '###'

#     metadata = {
#         "name": filename,
#         "parents": [folderid]
#     }
#     files = {
#         'data': ('metadata', json.dumps(metadata), 'application/json'),
#         'file': open(filedirectory, "rb").read()  # or  open(filedirectory, "rb")
#     }
#     r = requests.post(
#         "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart",
#         headers={"Authorization": "Bearer " + access_token},
#         files=files
#     )
#     print(r.text)


def sendEmail(request):
# Python code to illustrate Sending mail with attachments 
# from your Gmail account 

# libraries to be imported 
    import smtplib 
    from email.mime.multipart import MIMEMultipart 
    from email.mime.text import MIMEText 
    from email.mime.base import MIMEBase 
    from email import encoders 

    fromaddr = "saiyedayaz9@gmail.com"
    toaddr = "saiyedayaz9@gmail.com"

    # instance of MIMEMultipart 
    msg = MIMEMultipart() 

    # storing the senders email address 
    msg['From'] = fromaddr 

    # storing the receivers email address 
    msg['To'] = toaddr 

    # storing the subject 
    msg['Subject'] = "Customer Enquiry"

    # string to store the body of the mail 
    body = "Attached Excel Sheet"

    # attach the body with the msg instance 
    msg.attach(MIMEText(body, 'plain')) 

    # open the file to be sent 
    filename = "SoporteMoreliaCustomers.xlsx"
    attachment = open("./SoporteMoreliaCustomers.xlsx", "rb") 

    # instance of MIMEBase and named as p 
    p = MIMEBase('application', 'octet-stream') 

    # To change the payload into encoded form 
    p.set_payload((attachment).read()) 

    # encode into base64 
    encoders.encode_base64(p) 

    p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 

    # attach the instance 'p' to instance 'msg' 
    msg.attach(p) 

    # creates SMTP session 
    s = smtplib.SMTP('smtp.gmail.com', 587) 

    # start TLS for security 
    s.starttls() 

    # Authentication 
    s.login(fromaddr, "officialhearthackergmail7777") 

    # Converts the Multipart msg into a string 
    text = msg.as_string() 

    # sending the mail 
    s.sendmail(fromaddr, toaddr, text) 

    # terminating the session 
    s.quit()
    return


def temp(request):
    return render(request,'bot/yudizbot.html')


def option(request):
    return render(request,'bot/options.html')