#!/usr/bin/python
# -*- coding: latin-1 -*-

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

workbook = Workbook()
workbook = load_workbook(filename="SoporteMoreliaCustomers.xlsx")
sheet = workbook.active

# Let's create some sample sales data
rows = [
			[" Nombre del cliente "," Problema del cliente "," Correo electrónico "," fecha "]
       ]


for row in rows:
	sheet.append(row)

workbook.save(filename="SoporteMoreliaCustomers.xlsx")




# workbook.save(filename="CustomerLeads.xlsx")


# Create new file

# from openpyxl import Workbook

# workbook = Workbook()
# sheet = workbook.active
# workbook.save(filename="SoporteMoreliaCustomers.xlsx")

# # To auto fit cell texts
# for column_cells in sheet.columns:
#     length = 30
#     sheet.column_dimensions[column_cells[0].column_letter].width = length

# workbook.save(filename="SoporteMoreliaCustomers.xlsx")
